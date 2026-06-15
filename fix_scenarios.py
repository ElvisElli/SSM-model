#!/usr/bin/env python3
"""
Fix scenarios.csv by correcting crop_col mappings and re-reading crop parameters
for EU, NP (MG2->MG3, +4) and AL, NV (MG3->MG4, +4) locations.
"""

import csv
import openpyxl

XLSM_PATH = "/home/user/SSM-model/excel-model/SSM_iCrop_Soybean.xlsm"
SCENARIOS_PATH = "/home/user/SSM-model/r-model/inputs/scenarios.csv"

# Crop sheet row (1-based) -> CSV column name mapping
# CSV column names match the parameter names exactly (mixed case)
CROP_ROW_MAP = {
    5:  "crop_name",
    7:  "PHYL",
    8:  "PLACON",
    9:  "PLAPOW",
    10: "a_pla_den",
    11: "b_pla_den",
    12: "SLA",
    13: "FRZTKIL",
    14: "FRZLDR",
    15: "HtLTH",
    16: "HtLDR",
    17: "TBRUE",
    18: "TP1RUE",
    19: "TP2RUE",
    20: "TCRUE",
    21: "KPAR",
    22: "IRUE",
    # 23: skip (empty)
    24: "CO2REF",
    25: "CO2RES",
    26: "FLF1A",
    27: "FLF1B",
    28: "WTOPL",
    29: "FLF2",
    30: "FRTRL",
    31: "GCC",
    32: "PDHI",
    33: "WDHI1",
    34: "WDHI2",
    35: "WDHI3",
    36: "WDHI4",
    37: "MC",
    38: "heat",
    39: "TP2H",
    40: "TCH",
    41: "TP1F",
    42: "TBF",
    43: "RFmax",
    44: "DEPORT",
    45: "MEED",
    46: "GRTDP",
    47: "TECREF",
    48: "WSSG",
    49: "WSSL",
    50: "WSSD",
    51: "WSSN",
    52: "FLDKIL",
    53: "vpdtp",
    54: "VPDcr",
    55: "surviv",
    56: "EPCOND",
    57: "LTLRWC",
    58: "SLNG",
    59: "SLNS",
    60: "SNCG",
    61: "SNCS1",
    62: "SNCS2",
    63: "GNCmin",
    64: "GNCmax",
    65: "MXNUP",
    70: "TBD",
    71: "TP1D",
    72: "TP2D",
    73: "TCD",
    77: "cpp",
    78: "ppsen",
    80: "bdSOWEMR",
    81: "bdEMRR1",
    82: "bdR1R3",
    83: "bdR3R5",
    84: "bdR5R7",
    85: "bdR7R8",
    91: "bdBRP",
    92: "bdTRP",
    93: "bdBSG",
    94: "bdTSG",
    96: "bdTLM",
    97: "bdTLP",
    98: "bdBLS",
    99: "bdFLW",
    104: "bdBNF",
}

def get_prefix(scenario_name):
    """Extract 2-letter location prefix from scenario name."""
    return scenario_name[:2]

def get_crop_col_offset(prefix):
    """Return the offset to add to crop_col for the given prefix."""
    if prefix in ("EU", "NP"):
        return 4  # MG2 -> MG3
    elif prefix in ("AL", "NV"):
        return 4  # MG3 -> MG4
    else:
        return 0  # No change

def load_crop_sheet(xlsm_path):
    """Load the Crop sheet from the xlsm file."""
    print(f"Loading crop sheet from {xlsm_path}...")
    wb = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True)
    print(f"Available sheets: {wb.sheetnames}")
    crop_sheet = None
    for name in wb.sheetnames:
        if name.lower() == "crop":
            crop_sheet = wb[name]
            print(f"Found Crop sheet: '{name}'")
            break
    if crop_sheet is None:
        raise ValueError(f"No 'Crop' sheet found. Available: {wb.sheetnames}")
    return crop_sheet

def extract_crop_params(crop_sheet, col):
    """
    Extract all crop parameters from the given column (1-based).
    Returns a dict: csv_column_name -> value  (using exact case from CROP_ROW_MAP)
    """
    params = {}
    for row_num, param_name in CROP_ROW_MAP.items():
        cell = crop_sheet.cell(row=row_num, column=col)
        val = cell.value
        params[param_name] = val
    return params

def main():
    # Load crop sheet
    crop_sheet = load_crop_sheet(XLSM_PATH)

    # Cache extracted params by column number to avoid re-reading
    crop_cache = {}

    def get_crop_params(col):
        if col not in crop_cache:
            crop_cache[col] = extract_crop_params(crop_sheet, col)
        return crop_cache[col]

    # Read scenarios.csv
    print(f"\nReading {SCENARIOS_PATH}...")
    with open(SCENARIOS_PATH, "r", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)

    print(f"Loaded {len(rows)} rows with {len(fieldnames)} columns.")

    # Verify all crop param columns exist in CSV
    csv_cols = set(fieldnames)
    crop_param_names = set(CROP_ROW_MAP.values())
    missing = crop_param_names - csv_cols
    if missing:
        print(f"WARNING: These crop param columns are missing from CSV: {missing}")
    else:
        print("All crop parameter columns found in CSV header.")

    # Process rows
    changed_count = 0
    for row in rows:
        scenario = row["scenario"]
        prefix = get_prefix(scenario)
        offset = get_crop_col_offset(prefix)

        if offset != 0:
            old_col = int(row["crop_col"])
            new_col = old_col + offset
            row["crop_col"] = str(new_col)

            # Get crop params from new column
            new_params = get_crop_params(new_col)

            # Update all crop parameter columns in the row
            for param_name, val in new_params.items():
                if param_name in csv_cols:
                    row[param_name] = val if val is not None else ""
                else:
                    print(f"  WARNING: Column '{param_name}' not in CSV header, skipping.")

            changed_count += 1

    print(f"\nChanged {changed_count} rows.")

    # Verify key values before writing
    print("\n--- Verification ---")
    checks = {
        "AL-RFD-ELY-check": {"crop_col": "13", "cpp": 13.09, "ppsen": -0.294, "bdEMRR1": 19.7, "crop_name": "Soybean-MG4"},
        "EU-RFD-ELY-check": {"crop_col": "9", "cpp": 13.4, "ppsen": -0.285, "bdEMRR1": 19.5, "crop_name": "Soybean-MG3"},
        "JB-RFD-ELY-check": {"crop_col": "13"},
    }

    all_ok = True
    for scenario, expected in checks.items():
        row_dict = next((r for r in rows if r["scenario"] == scenario), None)
        if row_dict is None:
            print(f"  ERROR: Row '{scenario}' not found!")
            all_ok = False
            continue
        for field, exp_val in expected.items():
            actual = row_dict.get(field, "NOT_FOUND")
            # For numeric comparisons, try to convert
            try:
                actual_num = float(actual)
                exp_num = float(exp_val)
                match = abs(actual_num - exp_num) < 1e-6
            except (ValueError, TypeError):
                match = str(actual) == str(exp_val)
            status = "OK" if match else "FAIL"
            if not match:
                all_ok = False
            print(f"  {scenario} | {field}: expected={exp_val}, actual={actual} [{status}]")

    if not all_ok:
        print("\nWARNING: Verification failed! NOT writing file.")
        return

    print("\nAll verification checks passed!")

    # Write corrected CSV
    print(f"\nWriting corrected CSV to {SCENARIOS_PATH}...")
    with open(SCENARIOS_PATH, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print("Done writing.")

    # Summary: first row per location
    print("\n--- Summary: first row per location ---")
    print(f"{'Scenario':<30} {'crop_name':<30} {'cpp':<10} {'crop_col':<10}")
    seen_prefixes = set()
    for row in rows:
        scenario = row["scenario"]
        prefix = get_prefix(scenario)
        if prefix not in seen_prefixes:
            seen_prefixes.add(prefix)
            print(f"{scenario:<30} {row.get('crop_name',''):<30} {row.get('cpp',''):<10} {row.get('crop_col',''):<10}")

if __name__ == "__main__":
    main()
