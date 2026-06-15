#!/usr/bin/env python3
"""
Regenerate crop parameters in scenarios.csv using the CORRECT row numbers
from the Excel Crop sheet.

The previous fix_scenarios.py used wrong row numbers (off by 1-3 in several
places). This script re-reads ALL rows using the correct row-to-parameter
mapping verified from the actual Crop sheet layout.
"""

import csv
import openpyxl

XLSM_PATH = "/home/user/SSM-model/excel-model/SSM_iCrop_Soybean.xlsm"
SCENARIOS_PATH = "/home/user/SSM-model/r-model/inputs/scenarios.csv"

# CORRECT Crop sheet row (1-based) -> CSV column name mapping
# Verified against actual Crop sheet dump (June 2026)
CROP_ROW_MAP = {
    5:  "crop_name",
    7:  "PHYL",
    8:  "PLACON",
    9:  "PLAPOW",
    10: "a_pla_den",
    11: "b_pla_den",
    12: "SLA",
    13: "FRZTKIL",
    14: "FRZLDR",
    15: "HtLTH",
    16: "HtLDR",
    18: "TBRUE",     # was 17 (wrong)
    19: "TP1RUE",    # was 18 (wrong)
    20: "TP2RUE",    # was 19 (wrong)
    21: "TCRUE",     # was 20 (wrong)
    22: "KPAR",      # was 21 (wrong)
    23: "IRUE",      # was 22 (wrong); row 23 skip was incorrect
    24: "CO2REF",
    25: "CO2RES",
    # 26: empty (skip)
    27: "FLF1A",     # was 26 (wrong)
    28: "FLF1B",     # was 27 (wrong)
    29: "WTOPL",     # was 28 (wrong)
    30: "FLF2",      # was 29 (wrong)
    31: "FRTRL",     # was 30 (wrong)
    32: "GCC",       # was 31 (wrong); sheet label is "GCF"
    33: "PDHI",      # was 32 (wrong)
    34: "WDHI1",     # was 33 (wrong)
    35: "WDHI2",     # was 34 (wrong)
    36: "WDHI3",     # was 35 (wrong)
    37: "WDHI4",     # was 36 (wrong)
    38: "MC",        # was 37 (wrong)
    39: "heat",      # was 38 (wrong)
    40: "TP2H",      # was 39 (wrong)
    41: "TCH",       # was 40 (wrong)
    42: "TP1F",      # was 41 (wrong)
    43: "TBF",       # was 42 (wrong)
    44: "RFmax",     # was 43 (wrong)
    # 45: empty (skip)
    46: "DEPORT",    # was 44 (wrong); sheet label is "iDEPORT"
    47: "MEED",      # was 45 (wrong)
    48: "GRTDP",     # was 46 (wrong)
    49: "TECREF",    # was 47 (wrong); sheet label is "TEC"
    50: "WSSG",      # was 48 (wrong)
    51: "WSSL",      # was 49 (wrong)
    52: "WSSD",      # was 50 (wrong)
    53: "WSSN",      # was 51 (wrong)
    54: "FLDKIL",    # was 52 (wrong); sheet label is "FLDKL"
    55: "vpdtp",     # was 53 (wrong)
    56: "VPDcr",     # was 54 (wrong)
    57: "surviv",    # was 55 (wrong)
    58: "EPCOND",    # was 56 (wrong)
    59: "LTLRWC",    # was 57 (wrong)
    # 60: empty (skip)
    61: "SLNG",      # was 58 (wrong)
    62: "SLNS",      # was 59 (wrong)
    63: "SNCG",      # was 60 (wrong)
    64: "SNCS1",     # was 61 (wrong)
    65: "SNCS2",     # was 62 (wrong)
    66: "GNCmin",    # was 63 (wrong)
    67: "GNCmax",    # was 64 (wrong)
    68: "MXNUP",     # was 65 (wrong)
    70: "TBD",
    71: "TP1D",
    72: "TP2D",
    73: "TCD",
    77: "cpp",
    78: "ppsen",
    80: "bdSOWEMR",
    81: "bdEMRR1",
    82: "bdR1R3",
    83: "bdR3R5",
    84: "bdR5R7",
    85: "bdR7R8",
    91: "bdBRP",
    92: "bdTRP",
    93: "bdBSG",
    94: "bdTSG",
    96: "bdTLM",
    97: "bdTLP",
    98: "bdBLS",
    99: "bdFLW",
    104: "bdBNF",
}


def load_crop_sheet(xlsm_path):
    print(f"Loading crop sheet from {xlsm_path}...")
    wb = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True)
    crop_sheet = None
    for name in wb.sheetnames:
        if name.lower() == "crop":
            crop_sheet = wb[name]
            break
    if crop_sheet is None:
        raise ValueError(f"No 'Crop' sheet found. Available: {wb.sheetnames}")
    print(f"Found Crop sheet: '{name}'")
    return crop_sheet


def extract_crop_params(crop_sheet, col):
    """Extract crop parameters from column (1-based) using correct row mapping."""
    params = {}
    for row_num, param_name in CROP_ROW_MAP.items():
        cell = crop_sheet.cell(row=row_num, column=col)
        val = cell.value
        params[param_name] = val if val is not None else ""
    return params


def main():
    crop_sheet = load_crop_sheet(XLSM_PATH)

    # Cache by column number
    crop_cache = {}

    def get_crop_params(col):
        if col not in crop_cache:
            crop_cache[col] = extract_crop_params(crop_sheet, col)
        return crop_cache[col]

    # Verify key expected values from the crop sheet
    print("\n--- Verifying Crop sheet values ---")
    checks = [
        (13, "KPAR",   0.65),
        (13, "IRUE",   2),
        (13, "DEPORT", 200),
        (13, "MEED",   1000),
        (13, "TBRUE",  10),
        (13, "SLNG",   2.5),
        (13, "cpp",    13.09),
        (9,  "cpp",    13.4),
        (9,  "DEPORT", 200),
        (9,  "IRUE",   2),
    ]
    all_ok = True
    for col, param, expected in checks:
        params = get_crop_params(col)
        actual = params.get(param, "NOT_FOUND")
        try:
            match = abs(float(actual) - float(expected)) < 1e-4
        except (ValueError, TypeError):
            match = str(actual) == str(expected)
        status = "OK" if match else "FAIL"
        if not match:
            all_ok = False
        print(f"  col={col} {param}: expected={expected}, actual={actual} [{status}]")

    if not all_ok:
        print("\nERROR: Crop sheet verification failed. Aborting.")
        return

    print("\nAll verification checks passed!")

    # Read current scenarios.csv
    print(f"\nReading {SCENARIOS_PATH}...")
    with open(SCENARIOS_PATH, "r", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)
    print(f"Loaded {len(rows)} rows with {len(fieldnames)} columns.")

    # Verify crop param columns exist
    csv_cols = set(fieldnames)
    crop_param_names = set(CROP_ROW_MAP.values())
    missing = crop_param_names - csv_cols
    if missing:
        print(f"WARNING: Crop param columns missing from CSV: {missing}")

    # Re-read crop params for ALL rows using correct row numbers
    # crop_col in the CSV already has the correct MG column (13 for MG4, 9 for MG3, etc.)
    update_count = 0
    for row in rows:
        crop_col = int(row["crop_col"])
        new_params = get_crop_params(crop_col)
        for param_name, val in new_params.items():
            if param_name in csv_cols:
                row[param_name] = val if val is not None else ""
        update_count += 1

    print(f"Updated crop params in {update_count} rows.")

    # Final verification: spot-check key rows
    print("\n--- Final spot checks ---")
    spot_checks = {
        "JB-RFD-ELY-check": {"crop_col": "13", "KPAR": "0.65", "IRUE": "2",
                              "DEPORT": "200", "cpp": "13.09", "ppsen": "-0.294"},
        "AL-RFD-ELY-check": {"crop_col": "13", "KPAR": "0.65", "IRUE": "2",
                              "DEPORT": "200", "cpp": "13.09", "ppsen": "-0.294"},
        "EU-RFD-ELY-check": {"crop_col": "9",  "KPAR": "0.65", "IRUE": "2",
                              "DEPORT": "200", "cpp": "13.4",  "ppsen": "-0.285"},
        "NV-RFD-ELY-check": {"crop_col": "13", "KPAR": "0.65", "IRUE": "2",
                              "DEPORT": "200", "cpp": "13.09", "ppsen": "-0.294"},
    }
    all_spot_ok = True
    for scenario, expected in spot_checks.items():
        row_dict = next((r for r in rows if r["scenario"] == scenario), None)
        if row_dict is None:
            print(f"  ERROR: Row '{scenario}' not found!")
            all_spot_ok = False
            continue
        for field, exp_val in expected.items():
            actual = str(row_dict.get(field, "NOT_FOUND"))
            try:
                match = abs(float(actual) - float(exp_val)) < 1e-4
            except (ValueError, TypeError):
                match = actual == exp_val
            status = "OK" if match else "FAIL"
            if not match:
                all_spot_ok = False
            print(f"  {scenario} | {field}: expected={exp_val}, actual={actual} [{status}]")

    if not all_spot_ok:
        print("\nERROR: Spot checks failed. NOT writing output.")
        return

    # Write corrected CSV
    print(f"\nWriting corrected CSV to {SCENARIOS_PATH}...")
    with open(SCENARIOS_PATH, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print("Done!")

    # Summary: one row per location
    print("\n--- Summary: first row per location ---")
    print(f"{'Scenario':<30} {'crop_name':<22} {'crop_col':<9} {'cpp':<8} {'DEPORT':<8} {'IRUE'}")
    seen = set()
    for row in rows:
        scn = row["scenario"]
        prefix = scn[:2]
        if prefix not in seen:
            seen.add(prefix)
            print(f"{scn:<30} {row.get('crop_name',''):<22} {row.get('crop_col',''):<9} "
                  f"{row.get('cpp',''):<8} {row.get('DEPORT',''):<8} {row.get('IRUE','')}")


if __name__ == "__main__":
    main()
