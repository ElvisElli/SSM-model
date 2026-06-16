"""
Creates parameter_dictionary.xlsx — a reference spreadsheet describing every
column in scenarios.csv (management/scenario input file).
"""

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "parameter_dictionary.xlsx")

# ---------------------------------------------------------------------------
# Parameter definitions
# Format: (parameter, category, description, units, typical_range / notes)
# ---------------------------------------------------------------------------
PARAMS = [
    # ── Scenario identifiers ────────────────────────────────────────────────
    ("scenario",    "Scenario ID",  "Unique scenario name (location-management-cultivar key)",       "—",            "e.g. JB-RFD-ELY-check"),
    ("loc_row",     "Scenario ID",  "Row index of this location in the Excel Location sheet",        "integer",      "1–10"),
    ("mang_row",    "Scenario ID",  "Row index of this management in the Excel Management sheet",    "integer",      "1–24"),
    ("soil_row",    "Scenario ID",  "Row index of this soil profile in the Excel Soil sheet (also used as soil_data.json key)", "integer", "1–10"),
    ("crop_col",    "Scenario ID",  "Column index of this crop cultivar in the Excel Crop sheet",   "integer",      "13–16"),

    # ── Location / climate ──────────────────────────────────────────────────
    ("loc_name",    "Location",     "Full name of the simulation location",                          "—",            "e.g. Jonesboro"),
    ("lat",         "Location",     "Latitude of the location (used for daylength and hourly solar radiation)", "decimal degrees N", "30–43"),
    ("vpdf",        "Location",     "VPD scaling fraction: scales the VP gradient (VP(Tmax)−VP(Tmin)) to effective daytime VPD. Humid sites ≈ 0.75; drier sites > 0.75", "fraction", "0.60–0.90"),
    ("wth_file",    "Location",     "Weather file name (Excel .xlsx) in inputs/weather/",           "filename",     "SSM_Jonesboro_AR.xlsx"),
    ("tchng",       "Climate",      "Temperature offset added to all daily Tmin and Tmax (climate change scenario)", "°C", "0 (no change), ±1–4"),
    ("pchng",       "Climate",      "Precipitation multiplier applied to daily RAIN (1.0 = no change)", "fraction", "1.0 (no change)"),
    ("co2",         "Climate",      "Ambient CO₂ concentration for this simulation (used to adjust RUE and TEC)", "ppm", "380–500"),

    # ── Management ──────────────────────────────────────────────────────────
    ("manag_name",  "Management",   "Short name for the management treatment",                       "—",            "e.g. AR-RFD-ELY"),
    ("fix_find",    "Management",   "Sowing method: 0 = fixed date (pdoy), 1 = flexible (find sowing date by soil temp/moisture)", "0/1", "0"),
    ("fyear",       "Management",   "First simulation year (start of the multi-year loop)",          "year",         "1985–1995"),
    ("yrno",        "Management",   "Number of simulation years in the multi-year loop",             "years",        "30"),
    ("sim_doy",     "Management",   "Day of year when the simulation loop starts (weather reading begins here, before sowing)", "DOY", "115"),
    ("pdoy",        "Management",   "Planting (sowing) day of year",                                 "DOY",          "100–150 (early=~115, mid=~135, late=~150)"),
    ("lpdoy",       "Management",   "Last possible planting DOY (used when fix_find=1 for flexible sowing)", "DOY", "365"),
    ("stop_doy",    "Management",   "Force-stop DOY: simulation halted if maturity not reached by this date", "DOY", "365"),
    ("sow_tmp",     "Management",   "Minimum soil temperature (°C) required before sowing (used only when fix_find=1)", "°C", "10"),
    ("sow_wat",     "Management",   "Minimum soil water fraction (0–1) required before sowing (fix_find=1)", "fraction", "0.3"),
    ("pden",        "Management",   "Plant density (plants per m²); affects LAI expansion exponent", "plants/m²",   "20–45 (default 32)"),
    ("nitrogen",    "Management",   "Nitrogen simulation switch: 0 = no N limitation (disabled), 1 = N model active", "0/1", "0"),
    ("water",       "Management",   "Water balance mode: 1 = dryland (no irrigation), 2 = rainfed (soil water tracked), 3 = irrigated", "1/2/3", "2 or 3"),
    ("irglvl",      "Management",   "Irrigation trigger: fraction of transpirable soil water (FTSW) below which irrigation is applied (water=3 only)", "fraction", "0.5"),

    # ── Soil ────────────────────────────────────────────────────────────────
    ("soil_name",   "Soil",         "Descriptive name of the soil profile",                          "—",            "e.g. Soil_5"),
    ("nlyer",       "Soil",         "Number of soil layers (typically 10 for a 200-cm profile)",     "layers",       "10"),
    ("ldrain",      "Soil",         "Drainage layer depth — water drains freely below this layer",   "cm",           "—"),
    ("salb",        "Soil",         "Soil albedo (0–1): reflectance of bare soil surface",           "fraction",     "0.10–0.20"),
    ("u",           "Soil",         "Stage-1 soil evaporation limit (Upper limit): cumulative evaporation before transition to stage 2", "mm", "6–9"),
    ("cn2",         "Soil",         "SCS Curve Number (CN2) for runoff estimation at normal antecedent moisture", "—", "70–85"),

    # ── Crop: LAI and structure ─────────────────────────────────────────────
    ("crop_name",   "Crop",         "Cultivar name (identifies crop parameter set)",                 "—",            "e.g. Soybean-MG4"),
    ("PHYL",        "Crop – LAI",   "Phyllochron: biological days (BD) required to produce one leaf (or main-stem node)", "BD", "40–55"),
    ("PLACON",      "Crop – LAI",   "Leaf area expansion constant in the allometric function: PLA = (CBD/PHYL)^PLAPOW × PLACON", "cm²/plant", "1"),
    ("PLAPOW",      "Crop – LAI",   "Leaf area expansion exponent in the allometric power function (PLA ∝ CBD^PLAPOW)", "—", "2.0–3.0"),
    ("a_pla_den",   "Crop – LAI",   "Intercept of the plant-density correction for PLAPOW: PLAPOW_eff = PLAPOW × (a + b×PDEN)", "—", "~1.0"),
    ("b_pla_den",   "Crop – LAI",   "Slope of the plant-density correction for PLAPOW (negative: denser stands have smaller leaf area per plant)", "1/(plants/m²)", "~−0.0005"),
    ("SLA",         "Crop – LAI",   "Specific leaf area: leaf area per unit leaf dry mass. Controls conversion of leaf DM to LAI.", "m²/g", "0.025–0.040"),
    ("FRZTKIL",     "Crop – LAI",   "Frost kill temperature threshold: TMIN below this causes leaf area loss proportional to degree of freezing", "°C", "−2 to −5"),
    ("FRZLDR",      "Crop – LAI",   "Frost leaf damage rate: fraction of LAI lost per °C below FRZTKIL", "fraction/°C", "0.01–0.05"),
    ("HtLTH",       "Crop – LAI",   "Heat leaf-area threshold: TMAX above this causes leaf area loss",  "°C",         "40–45"),
    ("HtLDR",       "Crop – LAI",   "Heat leaf damage rate: fraction of LAI lost per °C above HtLTH", "fraction/°C", "0.01"),

    # ── DM Production ───────────────────────────────────────────────────────
    ("TBRUE",       "DM Production","Base temperature for RUE: below this temperature RUE = 0",     "°C",           "10"),
    ("TP1RUE",      "DM Production","Lower optimum temperature for RUE: RUE is maximum between TP1RUE and TP2RUE", "°C", "20"),
    ("TP2RUE",      "DM Production","Upper optimum temperature for RUE",                             "°C",           "30"),
    ("TCRUE",       "DM Production","Ceiling temperature for RUE: above this RUE = 0",              "°C",           "40"),
    ("KPAR",        "DM Production","PAR extinction coefficient in Beer's law (FINT = 1 − exp(−KPAR × LAI)). Controls canopy light interception.", "m²/m²", "0.5–0.7"),
    ("IRUE",        "DM Production","Intrinsic radiation use efficiency (base RUE at reference CO₂, optimal T, no stress): DDMP = SRAD × 0.48 × FINT × RUE", "g DM / MJ PAR", "1.5–2.5"),
    ("CO2REF",      "DM Production","Reference CO₂ concentration at which the cultivar parameters (IRUE, TECREF) were measured", "ppm", "350–420"),
    ("CO2RES",      "DM Production","CO₂ responsiveness coefficient: fractional increase in RUE per log10 unit CO₂ increase from 330 ppm", "fraction", "0.6–1.0"),
    ("vpdtp",       "DM Production","VPD integration mode: 0 = daily mean VPD (no LT trait); 1 = hourly integration (enables LT trait)", "0/1", "0 (check), 1 (LT cultivars)"),
    ("VPDcr",       "DM Production","Critical VPD threshold for the Limited Transpiration (LT) trait. Stomata constrain transpiration when hourly VPD > VPDcr. Set very high (9999) to disable.", "kPa", "1.5 / 2.0 / 2.5 (LT); 9999 (check)"),
    ("TECREF",      "DM Production","Transpiration efficiency coefficient at reference CO₂: TEC = TECREF × CO₂_factor. TR = DDMP × VPD / TEC", "kPa·g/m²/mm", "7–12"),
    ("surviv",      "DM Production","Crop survival fraction after extreme stress (not currently active in soybean simulations)", "fraction", "0"),
    ("EPCOND",      "DM Production","Epidermis conductance parameter (not currently used in active code)", "—", "0.1"),
    ("LTLRWC",      "DM Production","Leaf relative water content threshold for LT trait activation (not currently used)", "fraction", "0.55"),

    # ── DM Distribution ─────────────────────────────────────────────────────
    ("FLF1A",       "DM Distribution","Leaf fraction of new vegetative DM when total biomass (WTOP) < WTOPL (young, thin canopy)", "fraction", "0.55–0.70"),
    ("FLF1B",       "DM Distribution","Leaf fraction of new vegetative DM when WTOP ≥ WTOPL (established dense canopy)", "fraction", "0.35–0.50"),
    ("WTOPL",       "DM Distribution","WTOP threshold (g/m²) that switches leaf partitioning from FLF1A to FLF1B", "g/m²", "100–200"),
    ("FLF2",        "DM Distribution","Leaf fraction of new DM after bdTLM (terminal leaf maturity) — very small, mainly stems", "fraction", "0.01–0.10"),
    ("FRTRL",       "DM Distribution","Fraction of WTOP at beginning of seed growth (BSGDM) available for translocation to grain", "fraction", "0.15–0.30"),
    ("GCC",         "DM Distribution","Grain construction cost: g of structural DM per g of glucose consumed in grain synthesis", "g/g", "0.65–0.80"),
    ("PDHI",        "DM Distribution","Potential daily harvest index increment (g grain DM / g total DM / d) under no stress", "d⁻¹", "0.007–0.012"),
    ("WDHI1",       "DM Distribution","Lower WTOP threshold 1 for DHI modifier: DHIDMF = 0 when BSGDM < WDHI1", "g/m²", "0–100"),
    ("WDHI2",       "DM Distribution","Lower WTOP threshold 2: DHIDMF ramps up from 0→1 between WDHI1 and WDHI2", "g/m²", "100–200"),
    ("WDHI3",       "DM Distribution","Upper WTOP threshold 3: DHIDMF = 1 between WDHI2 and WDHI3 (optimal range)", "g/m²", "400–700"),
    ("WDHI4",       "DM Distribution","Upper WTOP threshold 4: DHIDMF declines from 1→0 above WDHI3; = 0 above WDHI4", "g/m²", "700–9999"),
    ("MC",          "DM Distribution","Grain moisture content at harvest; used to convert dry grain mass to wet yield: Ywet = WGRN / (1 − MC/100) × 10", "%", "13 (soybean standard)"),
    ("heat",        "DM Distribution","Heat/frost effect on DHI flag: 0 = disabled (no penalty), 1 = enabled", "0/1", "0"),
    ("TP2H",        "DM Distribution","Upper optimum temperature for DHI (heat effect threshold)",   "°C",           "—"),
    ("TCH",         "DM Distribution","Ceiling temperature for DHI heat effect",                     "°C",           "—"),
    ("TP1F",        "DM Production", "Lower optimum temperature for frost effect on DHI",            "°C",           "—"),
    ("TBF",         "DM Production", "Base temperature for frost effect on DHI",                     "°C",           "—"),
    ("RFmax",       "DM Production", "Maximum fractional reduction of DHI by frost/heat",            "fraction",     "—"),

    # ── Water stress thresholds ─────────────────────────────────────────────
    ("DEPORT",      "Soil Water",   "Maximum rooting depth explored by crop. Root front advances at GRTDP BD/day until DEPORT.", "cm", "150–200"),
    ("MEED",        "Soil Water",   "Minimum effective evaporation depth: depth of top soil layer available for stage-1 evaporation", "cm", "—"),
    ("GRTDP",       "Soil Water",   "Root depth growth rate: biological days per cm of root front advance", "BD/cm", "0.3–0.5"),
    ("WSSG",        "Soil Water",   "FTSW threshold for grain-fill water stress (WSFG): stress begins below this value", "fraction", "0.40–0.50"),
    ("WSSL",        "Soil Water",   "FTSW threshold for leaf-expansion water stress (WSFL): stress begins below this value", "fraction", "0.45–0.55"),
    ("WSSD",        "Soil Water",   "FTSW threshold for development water stress (WSFD): phenological development slows below this", "fraction", "0.40–0.50"),
    ("WSSN",        "Soil Water",   "FTSW threshold for N-fixation water stress (WSFN): N fixation is impaired below this", "fraction", "0.40–0.55"),
    ("FLDKIL",      "Soil Water",   "Number of consecutive flood days that kills the crop (sets MATYP=5)",  "days", "50"),

    # ── N stress thresholds (N module disabled — parameters reserved) ───────
    ("SLNG",        "Nitrogen",     "Stem N concentration — growth demand",                         "g N/g DM",    "—"),
    ("SLNS",        "Nitrogen",     "Stem N concentration — senescence supply",                     "g N/g DM",    "—"),
    ("SNCG",        "Nitrogen",     "Seed N concentration — growth demand",                         "g N/g DM",    "—"),
    ("SNCS1",       "Nitrogen",     "Seed N concentration — supply threshold 1",                    "g N/g DM",    "—"),
    ("SNCS2",       "Nitrogen",     "Seed N concentration — supply threshold 2",                    "g N/g DM",    "—"),
    ("GNCmin",      "Nitrogen",     "Minimum grain N concentration",                                 "g N/g DM",    "—"),
    ("GNCmax",      "Nitrogen",     "Maximum grain N concentration",                                 "g N/g DM",    "—"),
    ("MXNUP",       "Nitrogen",     "Maximum daily N uptake rate",                                   "g N/m²/d",    "—"),

    # ── Phenology cardinal temperatures ────────────────────────────────────
    ("TBD",         "Phenology",    "Base temperature for development (BD): biological days do not accumulate below TBD", "°C", "7 (soybean)"),
    ("TP1D",        "Phenology",    "Lower optimum temperature for development: maximum BD accumulation rate above TP1D", "°C", "26–28"),
    ("TP2D",        "Phenology",    "Upper optimum temperature: BD rate constant between TP1D and TP2D",  "°C",      "32–36"),
    ("TCD",         "Phenology",    "Ceiling temperature: no development above TCD",                  "°C",          "42–46"),
    ("cpp",         "Phenology",    "Critical photoperiod: soybean development is slowed when daylength exceeds cpp (short-day plant)", "h", "12.5–13.5"),
    ("ppsen",       "Phenology",    "Photoperiod sensitivity coefficient. Negative for short-day plants (soybean): ppfun = max(0, 1 + ppsen × (pp − cpp))", "h⁻¹", "−0.20 to −0.40"),

    # ── Phenological stage thresholds (cumulative BD) ───────────────────────
    ("bdSOWEMR",    "Phenology",    "BD from sowing to emergence (VE)",                              "BD",           "3–5"),
    ("bdEMRR1",     "Phenology",    "BD from emergence to R1 (beginning bloom)",                     "BD",           "15–25"),
    ("bdR1R3",      "Phenology",    "BD from R1 to R3 (beginning pod)",                              "BD",           "5–10"),
    ("bdR3R5",      "Phenology",    "BD from R3 to R5 (beginning seed fill)",                        "BD",           "5–12"),
    ("bdR5R7",      "Phenology",    "BD from R5 to R7 (physiological maturity)",                     "BD",           "25–40"),
    ("bdR7R8",      "Phenology",    "BD from R7 to R8 (harvest maturity)",                           "BD",           "10–15"),
    ("bdBRP",       "Phenology",    "BD at which photoperiod response begins (= bdSOWEMR, i.e., emergence)", "BD",   "4"),
    ("bdTRP",       "Phenology",    "BD at which photoperiod response ends (= bdSOWEMR + bdEMRR1 = R1)", "BD",      "23–27"),
    ("bdBSG",       "Phenology",    "BD at beginning of seed growth (BSG = R5 threshold)",           "BD",           "35–50"),
    ("bdTSG",       "Phenology",    "BD at end of seed growth (TSG = R7 threshold = physiological maturity)", "BD",  "65–90"),
    ("bdTLM",       "Phenology",    "BD at terminal leaf maturity (TLM): leaf partitioning switches from FLF1 to FLF2", "BD", "30–45"),
    ("bdTLP",       "Phenology",    "BD at terminal leaf pool (TLP): further leaf area senescence transition", "BD", "35–50"),
    ("bdBLS",       "Phenology",    "BD at beginning of leaf senescence (BLS): LAI starts declining toward maturity", "BD", "40–60"),
    ("bdFLW",       "Phenology",    "BD at flowering (= bdSOWEMR + bdEMRR1 = R1 equivalent)",       "BD",           "22–27"),
    ("bdBNF",       "Phenology",    "BD at which biological N fixation begins",                      "BD",           "12"),
]

# ---------------------------------------------------------------------------
# Output column descriptions (for yearly and daily outputs)
# ---------------------------------------------------------------------------
YEARLY_OUTPUTS = [
    ("sName",    "Identifier", "Scenario name (matches scenarios.csv scenario column)"),
    ("Location", "Identifier", "Location full name"),
    ("Manag",    "Identifier", "Management name (manag_name from scenarios.csv)"),
    ("Soil",     "Identifier", "Soil name"),
    ("Crop",     "Identifier", "Crop/cultivar name"),
    ("Pyear",    "Identifier", "Planting (simulation) year"),
    ("Pdoy",     "Identifier", "Planting day of year"),
    ("dtEMR",    "Phenology",  "Days after planting to emergence (VE)"),
    ("R0",       "Phenology",  "Days to R0 (not standard for soybean; −9 if not reached)"),
    ("R1",       "Phenology",  "Days to R1 (beginning bloom)"),
    ("R2",       "Phenology",  "Not used for soybean (always −9)"),
    ("R3",       "Phenology",  "Days to R3 (beginning pod)"),
    ("R4",       "Phenology",  "Not used for soybean (always −9)"),
    ("R5",       "Phenology",  "Days to R5 (beginning seed fill = BSG)"),
    ("R6",       "Phenology",  "Not used for soybean (always −9)"),
    ("R7",       "Phenology",  "Days to R7 (physiological maturity)"),
    ("R8",       "Phenology",  "Days to R8 (harvest maturity)"),
    ("MSNN",     "Biomass",    "Maximum stem node number (main stem nodes at maturity)"),
    ("MXLAI",    "Biomass",    "Maximum leaf area index reached during the season (m²/m²)"),
    ("R5ANTDM",  "Biomass",    "Total above-ground dry mass at R5 (beginning seed fill) g/m²"),
    ("WTOP",     "Biomass",    "Total above-ground dry matter at maturity (g/m²)"),
    ("WGRN",     "Biomass",    "Grain dry matter at maturity (g/m²)"),
    ("HI",       "Biomass",    "Harvest index = WGRN/WTOP at maturity (fraction)"),
    ("Ywet",     "Yield",      "Wet grain yield at standard moisture content: WGRN / (1 − MC/100) × 10  (kg/ha)"),
    ("IPASW",    "Water",      "Initial plant-available soil water at sowing (mm)"),
    ("CRAIN",    "Water",      "Cumulative seasonal rainfall from sowing to maturity (mm)"),
    ("CIRGW",    "Water",      "Cumulative irrigation water applied (mm)"),
    ("IRGNO",    "Water",      "Number of irrigation events during the season"),
    ("ATSWSL",   "Water",      "Available transpirable soil water at maturity across all layers (mm)"),
    ("CRUNOF",   "Water",      "Cumulative surface runoff (mm)"),
    ("CE",       "Water",      "Cumulative soil evaporation from sowing to maturity (mm)"),
    ("CTR",      "Water",      "Cumulative transpiration from sowing to maturity (mm)"),
    ("CDRAIN",   "Water",      "Cumulative deep drainage below root zone (mm)"),
    ("ET",       "Water",      "Total evapotranspiration = CE + CTR (mm)"),
    ("EoverET",  "Water",      "Evaporation fraction = CE / ET (soil evaporation as a fraction of total ET)"),
    ("NLF–CNDNIT","Nitrogen",  "Nitrogen output columns (all zero — N module is disabled in current simulations)"),
    ("MATYP",    "Status",     "Maturity type code: 1 = normal maturity, 2 = premature (LAI < 0.05 during seed fill), 5 = flood kill"),
    ("SRAINT",   "Environment","Total rainfall from sowing to maturity (mm) — same as CRAIN but computed from weather loop"),
    ("MTMINT",   "Environment","Mean daily minimum temperature from sowing to maturity (°C)"),
    ("MTMAXT",   "Environment","Mean daily maximum temperature from sowing to maturity (°C)"),
    ("SSRADT",   "Environment","Total solar radiation from sowing to maturity (MJ/m²)"),
    ("SUMETT",   "Environment","Total evapotranspiration from sowing to maturity (mm)"),
    ("SRAIN2",   "Environment","Total rainfall from sowing to beginning seed growth (BSG) (mm)"),
    ("MTMIN2",   "Environment","Mean daily Tmin from sowing to BSG (°C)"),
    ("MTMAX2",   "Environment","Mean daily Tmax from sowing to BSG (°C)"),
    ("SSRAD2",   "Environment","Total solar radiation from sowing to BSG (MJ/m²)"),
    ("SUMET2",   "Environment","Total ET from sowing to BSG (mm)"),
    ("SRAIN3",   "Environment","Total rainfall from BSG to maturity (grain fill period) (mm)"),
    ("MTMIN3",   "Environment","Mean daily Tmin from BSG to maturity (°C)"),
    ("MTMAX3",   "Environment","Mean daily Tmax from BSG to maturity (°C)"),
    ("SSRAD3",   "Environment","Total solar radiation from BSG to maturity (MJ/m²)"),
    ("SUMET3",   "Environment","Total ET from BSG to maturity (mm)"),
]

DAILY_OUTPUTS = [
    ("sName",    "Identifier",  "Scenario name"),
    ("Location", "Identifier",  "Location name"),
    ("Manag",    "Identifier",  "Management name"),
    ("Soil",     "Identifier",  "Soil name"),
    ("Crop",     "Identifier",  "Crop/cultivar name"),
    ("Pyear",    "Identifier",  "Simulation year"),
    ("doy",      "Time",        "Day of year (1–365)"),
    ("DAP",      "Time",        "Days after planting"),
    ("TMP",      "Weather",     "Mean daily temperature = (TMAX + TMIN) / 2  (°C)"),
    ("DTU",      "Phenology",   "Daily thermal units accumulated for phenology (°C·d, VBA definition)"),
    ("CBD",      "Phenology",   "Cumulative biological days (physiological age since sowing)"),
    ("MSNN",     "LAI",         "Main-stem node number"),
    ("GLAI",     "LAI",         "Leaf area index gained today (m²/m²/d)"),
    ("DLAI",     "LAI",         "Leaf area index lost today by senescence or damage (m²/m²/d)"),
    ("LAI",      "LAI",         "Leaf area index (m²/m²)"),
    ("TCFRUE",   "DM Prod.",    "Temperature correction factor for RUE (0–1, trapezoidal function of TMP)"),
    ("FINT",     "DM Prod.",    "Fraction of PAR intercepted by canopy = 1 − exp(−KPAR × LAI)"),
    ("DDMP",     "DM Prod.",    "Daily dry matter production (g/m²/d)"),
    ("GLF",      "DM Dist.",    "Leaf DM gained today (g/m²/d)"),
    ("GST",      "DM Dist.",    "Stem DM gained today (g/m²/d)"),
    ("SGR",      "DM Dist.",    "Seed (grain) growth rate today (g/m²/d)"),
    ("WLF",      "DM Dist.",    "Cumulative leaf dry mass (g/m²)"),
    ("WST",      "DM Dist.",    "Cumulative stem dry mass (g/m²)"),
    ("WVEG",     "DM Dist.",    "Cumulative vegetative dry mass = WLF + WST (g/m²)"),
    ("WGRN",     "DM Dist.",    "Cumulative grain dry mass (g/m²)"),
    ("WTOP",     "DM Dist.",    "Cumulative total above-ground dry mass (g/m²)"),
    ("DEPORT",   "Soil Water",  "Current effective rooting depth (cm)"),
    ("RAIN",     "Soil Water",  "Daily rainfall (mm/d)"),
    ("IRGW",     "Soil Water",  "Irrigation water applied today (mm/d)"),
    ("RUNOF",    "Soil Water",  "Surface runoff today (mm/d)"),
    ("PET",      "Soil Water",  "Potential evapotranspiration (mm/d)"),
    ("SEVP",     "Soil Water",  "Actual soil evaporation today (mm/d)"),
    ("TR",       "Soil Water",  "Actual plant transpiration today (mm/d)"),
    ("DRAIN",    "Soil Water",  "Deep drainage below root zone today (mm/d)"),
    ("ATSWRZ",   "Soil Water",  "Available transpirable soil water in root zone (mm)"),
    ("FTSWRZ",   "Soil Water",  "Fraction of transpirable soil water in root zone (0–1); drives water stress"),
    ("CRAIN",    "Soil Water",  "Cumulative rainfall from sowing to today (mm)"),
    ("CIRGW",    "Soil Water",  "Cumulative irrigation to today (mm)"),
    ("IRGNO",    "Soil Water",  "Cumulative number of irrigation events to today"),
    ("CRUNOF",   "Soil Water",  "Cumulative runoff to today (mm)"),
    ("CE",       "Soil Water",  "Cumulative soil evaporation to today (mm)"),
    ("CTR",      "Soil Water",  "Cumulative transpiration to today (mm)"),
    ("CDRAIN",   "Soil Water",  "Cumulative drainage to today (mm)"),
    ("WSFL",     "Stress",      "Water stress factor for leaf expansion (0 = max stress, 1 = no stress)"),
    ("WSFG",     "Stress",      "Water stress factor for grain fill / DM production"),
    ("WSFD",     "Stress",      "Water stress factor for phenological development"),
]

# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()

# ── colour palette ──────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1A5276")   # dark blue
CAT_FILL   = {
    "Scenario ID":        PatternFill("solid", fgColor="D4E6F1"),
    "Location":           PatternFill("solid", fgColor="D5F5E3"),
    "Climate":            PatternFill("solid", fgColor="D5F5E3"),
    "Management":         PatternFill("solid", fgColor="FDEBD0"),
    "Soil":               PatternFill("solid", fgColor="F9EBEA"),
    "Crop":               PatternFill("solid", fgColor="F5EEF8"),
    "Crop – LAI":         PatternFill("solid", fgColor="F5EEF8"),
    "DM Production":      PatternFill("solid", fgColor="FEF9E7"),
    "DM Distribution":    PatternFill("solid", fgColor="EAFAF1"),
    "Soil Water":         PatternFill("solid", fgColor="EBF5FB"),
    "Nitrogen":           PatternFill("solid", fgColor="FDEDEC"),
    "Phenology":          PatternFill("solid", fgColor="F0F3F4"),
}
DEFAULT_FILL = PatternFill("solid", fgColor="FFFFFF")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font():  return Font(bold=True, color="FFFFFF", size=11)
def cat_font():  return Font(bold=True, size=10)
def body_font(): return Font(size=10)
def wrap_align(): return Alignment(wrap_text=True, vertical="top")

def write_header(ws, cols, row=1):
    for c, (col_name, width) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=col_name)
        cell.font = hdr_font()
        cell.fill = PatternFill("solid", fgColor="1A5276")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = width

def write_row(ws, values, row, fill=None, bold=False):
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = fill or DEFAULT_FILL
        cell.font = Font(bold=bold, size=10)
        cell.alignment = wrap_align()
        cell.border = BORDER

# ── Sheet 1: Scenario/Management Input Parameters ──────────────────────────
ws1 = wb.active
ws1.title = "Input Parameters"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A2"

COLS1 = [("Parameter", 18), ("Category", 16), ("Description", 65),
         ("Units", 22), ("Typical range / notes", 35)]
write_header(ws1, COLS1)

prev_cat = None
for r, (param, cat, desc, units, notes) in enumerate(PARAMS, 2):
    fill = CAT_FILL.get(cat, DEFAULT_FILL)
    write_row(ws1, [param, cat, desc, units, notes], r, fill=fill)
    # Bold the parameter name
    ws1.cell(row=r, column=1).font = Font(bold=True, size=10, name="Courier New")

ws1.row_dimensions[1].height = 28
for r in range(2, len(PARAMS) + 2):
    ws1.row_dimensions[r].height = 38

# ── Sheet 2: Yearly Output Columns ─────────────────────────────────────────
ws2 = wb.create_sheet("Yearly Outputs")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A2"

COLS2 = [("Column", 14), ("Group", 16), ("Description", 80)]
write_header(ws2, COLS2)

CAT_FILL2 = {
    "Identifier":  PatternFill("solid", fgColor="D4E6F1"),
    "Phenology":   PatternFill("solid", fgColor="F0F3F4"),
    "Biomass":     PatternFill("solid", fgColor="D5F5E3"),
    "Yield":       PatternFill("solid", fgColor="EAFAF1"),
    "Water":       PatternFill("solid", fgColor="EBF5FB"),
    "Nitrogen":    PatternFill("solid", fgColor="FDEDEC"),
    "Status":      PatternFill("solid", fgColor="FEF9E7"),
    "Environment": PatternFill("solid", fgColor="F5EEF8"),
}
for r, (col, grp, desc) in enumerate(YEARLY_OUTPUTS, 2):
    fill = CAT_FILL2.get(grp, DEFAULT_FILL)
    write_row(ws2, [col, grp, desc], r, fill=fill)
    ws2.cell(row=r, column=1).font = Font(bold=True, size=10, name="Courier New")
    ws2.row_dimensions[r].height = 30
ws2.row_dimensions[1].height = 28

# ── Sheet 3: Daily Output Columns ──────────────────────────────────────────
ws3 = wb.create_sheet("Daily Outputs")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A2"

COLS3 = [("Column", 14), ("Group", 14), ("Description", 80)]
write_header(ws3, COLS3)

for r, (col, grp, desc) in enumerate(DAILY_OUTPUTS, 2):
    fill = CAT_FILL2.get(grp, CAT_FILL.get(grp, DEFAULT_FILL))
    write_row(ws3, [col, grp, desc], r, fill=fill)
    ws3.cell(row=r, column=1).font = Font(bold=True, size=10, name="Courier New")
    ws3.row_dimensions[r].height = 28
ws3.row_dimensions[1].height = 28

# ── Sheet 4: Abbreviations / Glossary ──────────────────────────────────────
ws4 = wb.create_sheet("Glossary")
ws4.sheet_view.showGridLines = False

GLOSSARY = [
    ("BD",      "Biological day — dimensionless physiological time unit (0–1 per calendar day). Accumulates based on temperature and photoperiod responses."),
    ("CBD",     "Cumulative biological days since sowing. Used to trigger phenological stages."),
    ("FTSW",    "Fraction of transpirable soil water — main water status variable driving stress factors (0 = wilting point, 1 = field capacity)."),
    ("ATSWRZ",  "Available transpirable soil water in the root zone (mm) = Σ(WL_L − LL_L) over all layers within rooting depth."),
    ("DUL",     "Drained upper limit — soil water content at field capacity (cm³/cm³)."),
    ("LL",      "Lower limit — soil water content at permanent wilting point (cm³/cm³)."),
    ("DDMP",    "Daily dry matter production (g/m²/d) = SRAD × 0.48 × FINT × RUE."),
    ("FINT",    "Fraction of PAR intercepted by the canopy = 1 − exp(−KPAR × LAI)."),
    ("RUE",     "Radiation use efficiency = IRUE × f_T(T) × WSFG  (g DM / MJ PAR)."),
    ("PAR",     "Photosynthetically active radiation (400–700 nm). Assumed = 0.48 × SRAD."),
    ("SRAD",    "Daily solar radiation (MJ/m²/d)."),
    ("VPD",     "Vapour pressure deficit (kPa) — difference between saturation VP at air temperature and actual VP. Drives transpiration demand."),
    ("TEC",     "Transpiration efficiency coefficient (kPa·g DM / m² / mm). TR = DDMP × VPD / TEC."),
    ("LT trait","Limited transpiration trait: stomata partially close when hourly VPD > VPDcr, reducing water use but also carbon gain. Parameterised by VPDcr."),
    ("BSG",     "Beginning of seed growth = R5 phenological stage. Grain filling starts."),
    ("TSG",     "Terminal seed growth = R7 (physiological maturity). Grain filling ends."),
    ("DHI",     "Dynamic Harvest Index rate — daily increment in HI = PDHI × DHIDMF (d⁻¹)."),
    ("HI",      "Harvest index = WGRN / WTOP."),
    ("MATYP",   "Maturity type flag: 1 = normal, 2 = premature (low LAI during seed fill), 5 = flood kill."),
    ("DOY",     "Day of year (1–365 or 1–366 for leap years)."),
    ("MG",      "Maturity group — soybean cultivar classification (MG4 is typical for mid-South US; later MGs for shorter-day environments)."),
    ("rainfed", "Water regime: soil water tracked, no irrigation added. Crop relies solely on rainfall and stored soil moisture."),
    ("irrigated","Water regime: irrigation applied automatically whenever FTSW drops below the trigger level (irglvl)."),
    ("check",   "Baseline cultivar with vpdtp=0 (daily VPD mode): no LT trait. Used as the reference for LT trait comparisons."),
    ("LT1.5/2/2.5", "LT-trait cultivars with VPDcr = 1.5, 2.0, 2.5 kPa respectively. Use hourly VPD integration (vpdtp=1)."),
    ("ELY/MID/LTE", "Planting window: ELY = early, MID = mid-season, LTE = late planting date."),
]

ws4.column_dimensions["A"].width = 16
ws4.column_dimensions["B"].width = 95

hdr = ws4.cell(row=1, column=1, value="Term / Abbreviation")
hdr.font = hdr_font(); hdr.fill = PatternFill("solid", fgColor="1A5276")
hdr.alignment = Alignment(horizontal="center", vertical="center")
hdr2 = ws4.cell(row=1, column=2, value="Definition")
hdr2.font = hdr_font(); hdr2.fill = PatternFill("solid", fgColor="1A5276")
hdr2.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 24

alt = [PatternFill("solid", fgColor="F2F3F4"),
       PatternFill("solid", fgColor="FDFEFE")]
for r, (term, defn) in enumerate(GLOSSARY, 2):
    fill = alt[(r) % 2]
    ws4.cell(row=r, column=1, value=term).font = Font(bold=True, size=10)
    ws4.cell(row=r, column=1).fill = fill
    ws4.cell(row=r, column=1).alignment = wrap_align()
    ws4.cell(row=r, column=2, value=defn).font = body_font()
    ws4.cell(row=r, column=2).fill = fill
    ws4.cell(row=r, column=2).alignment = wrap_align()
    ws4.row_dimensions[r].height = 30

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"  Sheet 1 'Input Parameters':  {len(PARAMS)} rows")
print(f"  Sheet 2 'Yearly Outputs':    {len(YEARLY_OUTPUTS)} rows")
print(f"  Sheet 3 'Daily Outputs':     {len(DAILY_OUTPUTS)} rows")
print(f"  Sheet 4 'Glossary':          {len(GLOSSARY)} rows")
